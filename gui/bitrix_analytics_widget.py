"""
Виджет аналитики по экспортам из Битрикс24.

Функции:
- Загрузка LEAD.csv и DEAL.csv
- Фильтрация "наших" лидов
- Визуализация метрик
- Экспорт отчётов
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QGroupBox,
    QFileDialog,
    QMessageBox,
    QTextEdit,
    QScrollArea,
)
from PyQt6.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

from services.bitrix_analytics_service import BitrixAnalyticsService
from config.settings import settings


logger = logging.getLogger(__name__)


class BitrixAnalyticsWidget(QWidget):
    """Виджет аналитики по экспортам Битрикс24."""

    def __init__(self, parent=None):
        super().__init__(parent)

        # ========== ТЕМНАЯ ТЕМА ДЛЯ MATPLOTLIB ==========
        import matplotlib
        matplotlib.rcParams.update({
            'figure.facecolor': '#1e1e1e',      # Фон графика
            'axes.facecolor': '#2d2d2d',        # Фон области с данными
            'axes.edgecolor': '#555',           # Цвет рамки
            'axes.labelcolor': '#e0e0e0',       # Цвет подписей осей
            'text.color': '#e0e0e0',            # Цвет текста
            'xtick.color': '#e0e0e0',           # Цвет меток X
            'ytick.color': '#e0e0e0',           # Цвет меток Y
            'grid.color': '#444',               # Цвет сетки
            'legend.facecolor': '#2d2d2d',      # Фон легенды
            'legend.edgecolor': '#555',         # Рамка легенды
        })

        # Сервис
        self.analytics_service = BitrixAnalyticsService()

        # Пути к загруженным файлам
        self.lead_path: Optional[Path] = None
        self.deal_path: Optional[Path] = None

        self._setup_ui()
        logger.info("Виджет Битрикс-аналитики инициализирован")

    def _setup_ui(self):
        """Создать UI."""
        main_layout = QVBoxLayout()

        # Заголовок
        header = QLabel("📊 Битрикс24 — Аналитика результатов")
        header.setStyleSheet(
            "font-size: 18px; font-weight: bold; margin: 10px;")
        main_layout.addWidget(header)

        # Скролл-область
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        # Блок 1: Загрузка файлов
        load_group = self._create_load_section()
        scroll_layout.addWidget(load_group)

        # Блок 2: Статистика (карточки)
        self.stats_group = self._create_stats_section()
        scroll_layout.addWidget(self.stats_group)

        # Блок 3: Графики
        self.charts_group = self._create_charts_section()
        scroll_layout.addWidget(self.charts_group)

        # Блок 4: Текстовый отчёт
        self.report_group = self._create_report_section()
        scroll_layout.addWidget(self.report_group)

        # Блок 5: Экспорт
        export_group = self._create_export_section()
        scroll_layout.addWidget(export_group)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)

        # ========== ТЕМНАЯ ТЕМА ДЛЯ ВСЕГО ВИДЖЕТА ==========
        self.setStyleSheet("""
            QWidget {
                background-color: #1e1e1e;
                color: #e0e0e0;
            }
            QGroupBox {
                background-color: #2d2d2d;
                border: 1px solid #444;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QLabel {
                color: #e0e0e0;
                background: transparent;
            }
            QPushButton {
                background-color: #3a3a3a;
                color: #ffffff;
                border: 1px solid #555;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #4a4a4a;
                border: 1px solid #666;
            }
            QPushButton:pressed {
                background-color: #2a2a2a;
            }
            QPushButton:disabled {
                background-color: #2a2a2a;
                color: #666;
            }
            QTextEdit {
                background-color: #2d2d2d;
                color: #d4d4d4;
                border: 1px solid #444;
                border-radius: 5px;
                padding: 10px;
            }
            QScrollArea {
                background-color: #1e1e1e;
                border: none;
            }
        """)

        self.setLayout(main_layout)

    def _create_load_section(self) -> QGroupBox:
        """Блок загрузки файлов."""
        group = QGroupBox("1. Загрузка экспортов из Битрикс24")
        layout = QVBoxLayout()

        # Инструкция
        instruction = QLabel(
            "💡 Инструкция:\n"
            "1. Экспортируйте из Битрикс24 два файла:\n"
            "   • LEAD.csv — все лиды (Некондиция)\n"
            "   • DEAL.csv — все сделки (В работе)\n"
            "2. Загрузите их ниже для анализа"
        )
        instruction.setStyleSheet(
            "color: #666; padding: 10px; background: #f8f9fa; border-radius: 5px;")
        layout.addWidget(instruction)

        # Кнопки загрузки
        buttons_layout = QHBoxLayout()

        self.btn_load_lead = QPushButton("📄 Загрузить LEAD.csv")
        self.btn_load_lead.clicked.connect(self._on_load_lead_clicked)

        self.btn_load_deal = QPushButton("📄 Загрузить DEAL.csv")
        self.btn_load_deal.clicked.connect(self._on_load_deal_clicked)

        self.btn_analyze = QPushButton("🔍 Анализировать")
        self.btn_analyze.clicked.connect(self._on_analyze_clicked)
        self.btn_analyze.setEnabled(False)
        self.btn_analyze.setStyleSheet(
            "background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;"
        )

        buttons_layout.addWidget(self.btn_load_lead)
        buttons_layout.addWidget(self.btn_load_deal)
        buttons_layout.addWidget(self.btn_analyze)

        layout.addLayout(buttons_layout)

        # Статус загрузки
        self.label_load_status = QLabel("⏳ Загрузите оба файла")
        self.label_load_status.setStyleSheet(
            "color: #999; font-style: italic;")
        layout.addWidget(self.label_load_status)

        group.setLayout(layout)
        return group

    def _create_stats_section(self) -> QGroupBox:
        """Блок статистики (карточки)."""
        group = QGroupBox("2. Общая статистика")
        layout = QVBoxLayout()

        # Карточки создаются при анализе
        self.stats_layout = QHBoxLayout()
        layout.addLayout(self.stats_layout)

        group.setLayout(layout)
        group.setVisible(False)  # Скрыто до анализа
        return group

    def _create_charts_section(self) -> QGroupBox:
        """Блок графиков."""
        group = QGroupBox("3. Визуализация")
        layout = QVBoxLayout()

        # График 1: Причины отказа (круговая диаграмма)
        layout.addWidget(QLabel("Причины отказа (ТОП-5):"))
        self.chart_rejections = FigureCanvas(
            Figure(figsize=(12, 8)))  # ← было (8, 6)
        self.chart_rejections.setMinimumHeight(500)  # ← ДОБАВЬ ЭТУ СТРОКУ
        layout.addWidget(self.chart_rejections)

        # График 2: Стадии сделок (столбчатая диаграмма)
        layout.addWidget(QLabel("Стадии сделок:"))
        self.chart_stages = FigureCanvas(
            Figure(figsize=(14, 6)))  # ← было (10, 5)
        self.chart_stages.setMinimumHeight(400)  # ← ДОБАВЬ ЭТУ СТРОКУ
        layout.addWidget(self.chart_stages)

        # График 3: Топ-менеджеры (столбчатая диаграмма)
        layout.addWidget(QLabel("Топ-менеджеры по количеству сделок:"))
        self.chart_managers = FigureCanvas(
            Figure(figsize=(14, 6)))  # ← было (10, 5)
        self.chart_managers.setMinimumHeight(400)  # ← ДОБАВЬ ЭТУ СТРОКУ
        layout.addWidget(self.chart_managers)

        group.setLayout(layout)
        group.setVisible(False)
        return group

    def _create_report_section(self) -> QGroupBox:
        """Блок текстового отчёта."""
        group = QGroupBox("4. Текстовая сводка")
        layout = QVBoxLayout()

        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        self.report_text.setStyleSheet(
            """
            QTextEdit {
                font-family: 'Courier New', monospace; 
                background: #1e1e1e;
                color: #d4d4d4;
                padding: 10px;
                border: 1px solid #444;
                border-radius: 5px;
            }
            """
        )
        self.report_text.setMinimumHeight(300)
        layout.addWidget(self.report_text)

        group.setLayout(layout)
        group.setVisible(False)
        return group

    def _create_export_section(self) -> QGroupBox:
        """Блок экспорта."""
        group = QGroupBox("5. Экспорт отчёта")
        layout = QHBoxLayout()

        self.btn_export_txt = QPushButton("📄 Экспорт в TXT")
        self.btn_export_txt.clicked.connect(self._on_export_txt_clicked)
        self.btn_export_txt.setEnabled(False)

        self.btn_export_excel = QPushButton("📊 Экспорт в Excel")
        self.btn_export_excel.clicked.connect(self._on_export_excel_clicked)
        self.btn_export_excel.setEnabled(False)

        layout.addWidget(self.btn_export_txt)
        layout.addWidget(self.btn_export_excel)
        layout.addStretch()

        group.setLayout(layout)
        return group

    def _on_load_lead_clicked(self):
        """Загрузить LEAD.csv."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите LEAD.csv",
            "",
            "CSV Files (*.csv);;All Files (*)",
        )

        if file_path:
            self.lead_path = Path(file_path)
            logger.info(f"LEAD выбран: {self.lead_path.name}")
            self._update_load_status()

    def _on_load_deal_clicked(self):
        """Загрузить DEAL.csv."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите DEAL.csv",
            "",
            "CSV Files (*.csv);;All Files (*)",
        )

        if file_path:
            self.deal_path = Path(file_path)
            logger.info(f"DEAL выбран: {self.deal_path.name}")
            self._update_load_status()

    def _update_load_status(self):
        """Обновить статус загрузки."""
        if self.lead_path and self.deal_path:
            self.label_load_status.setText(
                f"✅ LEAD: {self.lead_path.name}\n"
                f"✅ DEAL: {self.deal_path.name}"
            )
            self.label_load_status.setStyleSheet("color: green;")
            self.btn_analyze.setEnabled(True)
        elif self.lead_path:
            self.label_load_status.setText(
                f"✅ LEAD: {self.lead_path.name}\n⏳ Загрузите DEAL.csv")
            self.label_load_status.setStyleSheet("color: orange;")
        elif self.deal_path:
            self.label_load_status.setText(
                f"⏳ Загрузите LEAD.csv\n✅ DEAL: {self.deal_path.name}")
            self.label_load_status.setStyleSheet("color: orange;")

    def _on_analyze_clicked(self):
        """Запустить анализ."""
        if not self.lead_path or not self.deal_path:
            QMessageBox.warning(self, "Ошибка", "Загрузите оба файла!")
            return

        try:
            # Шаг 1: Загрузка файлов
            self.label_load_status.setText("⏳ Загрузка файлов...")
            success, message = self.analytics_service.load_bitrix_exports(
                self.lead_path, self.deal_path
            )

            if not success:
                QMessageBox.critical(self, "Ошибка загрузки", message)
                return

            # Шаг 2: Фильтрация "наших" лидов
            self.label_load_status.setText("⏳ Фильтрация 'наших' лидов...")
            total_before, total_after = self.analytics_service.filter_my_leads()

            logger.info(f"Фильтрация: {total_before} → {total_after}")

            # Шаг 3: Расчёт метрик
            self.label_load_status.setText("⏳ Расчёт метрик...")
            metrics = self.analytics_service.calculate_metrics()

            # Шаг 4: Отображение результатов
            self._display_results(metrics)

            self.label_load_status.setText(
                f"✅ Анализ завершён!\n"
                f"Всего записей: {total_before} → Наших лидов: {total_after}"
            )
            self.label_load_status.setStyleSheet(
                "color: green; font-weight: bold;")

            # Включаем кнопки экспорта
            self.btn_export_txt.setEnabled(True)
            self.btn_export_excel.setEnabled(True)

            logger.info("Анализ завершён успешно")

        except Exception as exc:
            logger.exception("Ошибка при анализе")
            QMessageBox.critical(
                self,
                "Ошибка анализа",
                f"Не удалось выполнить анализ:\n{exc}",
            )

    def _display_results(self, metrics: dict):
        """Отобразить результаты анализа."""
        # Показываем все блоки
        self.stats_group.setVisible(True)
        self.charts_group.setVisible(True)
        self.report_group.setVisible(True)

        # Карточки статистики
        self._create_stat_cards(metrics)

        # Графики
        self._create_rejection_chart(metrics)
        self._create_stages_chart(metrics)
        self._create_managers_chart(metrics)

        # Текстовый отчёт ← ПРОВЕРЬ ЭТУ ЧАСТЬ
        report_text = self.analytics_service.get_report_summary()
        self.report_text.setPlainText(report_text)  # ← ЭТА СТРОКА ДОЛЖНА БЫТЬ!

        # ← ДОБАВЬ ОТЛАДКУ:
        logger.info(f"Длина текста отчёта: {len(report_text)} символов")

    def _create_stat_cards(self, metrics: dict):
        """Создать карточки статистики."""
        # Очищаем старые карточки
        while self.stats_layout.count():
            child = self.stats_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Создаём новые
        cards_data = [
            ("Всего записей", f"{metrics.get('total_leads', 0):,}"),
            ("В работе (DEAL)", f"{metrics.get('total_deal_records', 0):,}"),
            ("Отказы (LEAD)", f"{metrics.get('total_rejections', 0):,}"),
            ("Успешные продажи", f"{metrics.get('successful_deals', 0):,}"),
            ("Конверсия", f"{metrics.get('conversion', 0)}%"),
        ]

        for title, value in cards_data:
            card = self._create_stat_card(title, value)
            self.stats_layout.addWidget(card)

    def _create_stat_card(self, title: str, value: str) -> QWidget:
        """Создать одну карточку статистики (темная тема)."""
        card = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)

        # Заголовок
        title_label = QLabel(title)
        title_label.setStyleSheet(
            "font-size: 12px; color: #999; font-weight: normal;"
        )
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Значение
        value_label = QLabel(value)
        value_label.setStyleSheet(
            "font-size: 24px; color: #64B5F6; font-weight: bold;"  # Голубой цвет
        )
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        card.setLayout(layout)
        card.setStyleSheet(
            """
            QWidget {
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2d2d2d, stop:1 #252525
                );
                border-radius: 8px;
                border: 1px solid #444;
            }
            """
        )
        card.setMinimumHeight(100)
        card.setMinimumWidth(150)

        return card

    def _create_rejection_chart(self, metrics: dict):
        """Создать график причин отказа (темная тема)."""
        rejection_reasons = metrics.get("rejection_reasons", {})

        fig = self.chart_rejections.figure
        fig.clear()
        ax = fig.add_subplot(111)

        if not rejection_reasons:
            ax.text(
                0.5, 0.5, "Нет данных о причинах отказа",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
        else:
            # ТОП-5 причин
            top_reasons = dict(sorted(
                rejection_reasons.items(),
                key=lambda x: x[1],
                reverse=True
            )[:5])

            labels = list(top_reasons.keys())
            sizes = list(top_reasons.values())

            # Темная цветовая палитра
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8']

            wedges, texts, autotexts = ax.pie(
                sizes,
                labels=labels,
                autopct="%1.1f%%",
                startangle=90,
                colors=colors,
                textprops={'color': '#e0e0e0'}
            )

            # Белый цвет для процентов
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_weight('bold')

            ax.set_title(
                "Причины отказа (ТОП-5)",
                fontsize=14,
                fontweight="bold",
                color='#ffffff',
                pad=20
            )

        fig.tight_layout()
        self.chart_rejections.draw()

    def _create_stages_chart(self, metrics: dict):
        """Создать график стадий сделок (темная тема)."""
        deal_stages = metrics.get("deal_stages", {})

        fig = self.chart_stages.figure
        fig.clear()
        ax = fig.add_subplot(111)

        if not deal_stages:
            ax.text(
                0.5, 0.5, "Нет данных о стадиях сделок",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
        else:
            stages = list(deal_stages.keys())
            counts = list(deal_stages.values())

            bars = ax.bar(
                range(len(stages)),
                counts,
                color='#4ECDC4',  # Бирюзовый
                edgecolor='#2d9cdb',
                linewidth=1.5
            )

            # Градиент для баров
            for i, bar in enumerate(bars):
                bar.set_color(['#4ECDC4', '#45B7D1', '#5DADE2',
                              '#64B5F6', '#7EC8E3'][i % 5])

            ax.set_xticks(range(len(stages)))
            ax.set_xticklabels(stages, rotation=45, ha="right")
            ax.set_xlabel("Стадия", fontsize=12, color='#e0e0e0')
            ax.set_ylabel("Количество сделок", fontsize=12, color='#e0e0e0')
            ax.set_title(
                "Стадии сделок",
                fontsize=14,
                fontweight="bold",
                color='#ffffff',
                pad=20
            )
            ax.grid(axis="y", alpha=0.2, color='#555')

        fig.tight_layout()
        self.chart_stages.draw()

    def _create_managers_chart(self, metrics: dict):
        """Создать график топ-менеджеров (темная тема)."""
        top_managers = metrics.get("top_managers", {})

        fig = self.chart_managers.figure
        fig.clear()
        ax = fig.add_subplot(111)

        if not top_managers:
            ax.text(
                0.5, 0.5, "Нет данных о менеджерах",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
        else:
            managers = list(top_managers.keys())
            counts = list(top_managers.values())

            bars = ax.bar(
                range(len(managers)),
                counts,
                color='#FF6B6B',  # Коралловый
                edgecolor='#e55353',
                linewidth=1.5
            )

            # Градиент
            for i, bar in enumerate(bars):
                bar.set_color(['#FF6B6B', '#FF8787', '#FFA07A',
                              '#FFB399', '#FFC4B3'][i % 5])

            ax.set_xticks(range(len(managers)))
            ax.set_xticklabels(managers, rotation=45, ha="right")
            ax.set_xlabel("Менеджер", fontsize=12, color='#e0e0e0')
            ax.set_ylabel("Количество сделок", fontsize=12, color='#e0e0e0')
            ax.set_title(
                "Топ-менеджеры",
                fontsize=14,
                fontweight="bold",
                color='#ffffff',
                pad=20
            )
            ax.grid(axis="y", alpha=0.2, color='#555')

        fig.tight_layout()
        self.chart_managers.draw()

    def _on_export_txt_clicked(self):
        """Экспортировать отчёт в TXT."""
        try:
            default_name = f"bitrix_report_{pd.Timestamp.now().strftime('%Y-%m-%d_%H-%M-%S')}.txt"
            default_path = str(settings.paths.reports_dir / default_name)

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить отчёт TXT",
                default_path,
                "Text Files (*.txt);;All Files (*)",
            )

            if not file_path:
                return

            output_path = Path(file_path)

            # Получаем текстовую сводку
            report_text = self.analytics_service.get_report_summary()

            # Сохраняем в файл
            output_path.write_text(report_text, encoding="utf-8")

            QMessageBox.information(
                self,
                "✅ Экспорт завершён",
                f"Отчёт сохранён:\n{output_path}",
            )

            logger.info(f"TXT-отчёт экспортирован: {output_path}")

        except Exception as exc:
            logger.exception("Ошибка экспорта TXT")
            QMessageBox.critical(
                self,
                "❌ Ошибка",
                f"Не удалось экспортировать отчёт:\n{exc}",
            )

    def _on_export_excel_clicked(self):
        """Экспортировать отчёт в Excel."""
        try:
            default_name = f"bitrix_report_{pd.Timestamp.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            default_path = str(settings.paths.reports_dir / default_name)

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить отчёт Excel",
                default_path,
                "Excel Files (*.xlsx);;All Files (*)",
            )

            if not file_path:
                return

            output_path = Path(file_path)

            # Создаём Excel с несколькими листами
            self._export_to_excel(output_path)

            QMessageBox.information(
                self,
                "✅ Экспорт завершён",
                f"Отчёт сохранён:\n{output_path}",
            )

            logger.info(f"Excel-отчёт экспортирован: {output_path}")

        except Exception as exc:
            logger.exception("Ошибка экспорта Excel")
            QMessageBox.critical(
                self,
                "❌ Ошибка",
                f"Не удалось экспортировать отчёт:\n{exc}",
            )

    def _export_to_excel(self, output_path: Path):
        """
        Экспортировать детальный отчёт в Excel.

        Создаёт несколько листов:
        - Сводка
        - Причины отказа
        - Стадии сделок
        - Топ-менеджеры
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import PieChart, BarChart, Reference

        wb = Workbook()

        # Лист 1: СВОДКА
        ws_summary = wb.active
        ws_summary.title = "Сводка"

        # Заголовок
        ws_summary["A1"] = "ОТЧЁТ ПО БИТРИКС24"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary.merge_cells("A1:B1")

        # Дата
        ws_summary["A2"] = "Дата генерации:"
        ws_summary["B2"] = pd.Timestamp.now().strftime("%d.%m.%Y %H:%M")

        # Метрики
        metrics = self.analytics_service.metrics
        row = 4

        summary_data = [
            ("Всего записей", metrics.get("total_leads", 0)),
            ("Лиды (LEAD)", metrics.get("total_lead_records", 0)),
            ("Сделки (DEAL)", metrics.get("total_deal_records", 0)),
            ("Отказы", metrics.get("total_rejections", 0)),
            ("Успешные продажи", metrics.get("successful_deals", 0)),
            ("Конверсия", f"{metrics.get('conversion', 0)}%"),
        ]

        for label, value in summary_data:
            ws_summary[f"A{row}"] = label
            ws_summary[f"B{row}"] = value
            ws_summary[f"A{row}"].font = Font(bold=True)
            row += 1

        # Лист 2: ПРИЧИНЫ ОТКАЗА
        ws_rejections = wb.create_sheet("Причины отказа")
        ws_rejections["A1"] = "Причина"
        ws_rejections["B1"] = "Количество"
        ws_rejections["C1"] = "Процент"

        # Стили заголовка
        for cell in ["A1", "B1", "C1"]:
            ws_rejections[cell].font = Font(bold=True)
            ws_rejections[cell].fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            ws_rejections[cell].font = Font(color="FFFFFF", bold=True)

        rejection_reasons = metrics.get("rejection_reasons", {})
        total_rej = metrics.get("total_rejections", 1)

        row = 2
        for reason, count in sorted(
            rejection_reasons.items(), key=lambda x: x[1], reverse=True
        ):
            ws_rejections[f"A{row}"] = reason
            ws_rejections[f"B{row}"] = count
            ws_rejections[f"C{row}"] = f"{(count / total_rej * 100):.1f}%"
            row += 1

        # Лист 3: СТАДИИ СДЕЛОК
        ws_stages = wb.create_sheet("Стадии сделок")
        ws_stages["A1"] = "Стадия"
        ws_stages["B1"] = "Количество"

        # Стили
        for cell in ["A1", "B1"]:
            ws_stages[cell].font = Font(bold=True)
            ws_stages[cell].fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
            ws_stages[cell].font = Font(color="FFFFFF", bold=True)

        deal_stages = metrics.get("deal_stages", {})
        row = 2
        for stage, count in sorted(deal_stages.items(), key=lambda x: x[1], reverse=True):
            ws_stages[f"A{row}"] = stage
            ws_stages[f"B{row}"] = count
            row += 1

        # Лист 4: ТОП-МЕНЕДЖЕРЫ
        ws_managers = wb.create_sheet("Топ-менеджеры")
        ws_managers["A1"] = "Менеджер"
        ws_managers["B1"] = "Количество сделок"

        # Стили
        for cell in ["A1", "B1"]:
            ws_managers[cell].font = Font(bold=True)
            ws_managers[cell].fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
            ws_managers[cell].font = Font(color="000000", bold=True)

        top_managers = metrics.get("top_managers", {})
        row = 2
        for manager, count in top_managers.items():
            ws_managers[f"A{row}"] = manager
            ws_managers[f"B{row}"] = count
            row += 1

        # Автоширина колонок для всех листов
        for ws in wb.worksheets:
            for col in ["A", "B", "C"]:
                ws.column_dimensions[col].width = 30

        # Сохраняем
        wb.save(output_path)
        logger.info(f"Excel-отчёт с {len(wb.worksheets)} листами сохранён")
