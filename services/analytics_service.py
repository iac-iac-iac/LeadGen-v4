"""
Сервис аналитики и статистики.

Генерирует графики, отчёты и метрики по обработанным данным.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure

from repositories.processing_history_repo import ProcessingHistoryRepository
from repositories.managers_repo import ManagersRepository


logger = logging.getLogger(__name__)


class AnalyticsService:
    """Сервис аналитики и визуализации данных."""

    def __init__(
        self,
        history_repo: ProcessingHistoryRepository,
        managers_repo: ManagersRepository,
    ):
        self.history_repo = history_repo
        self.managers_repo = managers_repo

        # Настройка matplotlib для кириллицы
        plt.rcParams["font.family"] = "DejaVu Sans"
        plt.rcParams["axes.unicode_minus"] = False

    def get_overall_stats(self) -> Dict[str, int]:
        """
        Получить общую статистику по всем обработкам.

        Returns:
            Словарь с метриками.
        """
        query = """
        SELECT 
            COUNT(*) as total_processings,
            SUM(file_count) as total_files,
            SUM(total_rows) as total_rows,
            SUM(final_rows) as total_valid_leads,
            SUM(removed_duplicates) as total_duplicates,
            SUM(removed_empty_phones) as total_invalid_phones,
            SUM(unique_phones) as total_unique_phones
        FROM processing_history
        WHERE status = 'success'
        """

        row = self.history_repo.fetch_one(query)

        if not row:
            return {
                "total_processings": 0,
                "total_files": 0,
                "total_rows": 0,
                "total_valid_leads": 0,
                "total_duplicates": 0,
                "total_invalid_phones": 0,
                "total_unique_phones": 0,
            }

        return {
            "total_processings": row["total_processings"] or 0,
            "total_files": row["total_files"] or 0,
            "total_rows": row["total_rows"] or 0,
            "total_valid_leads": row["total_valid_leads"] or 0,
            "total_duplicates": row["total_duplicates"] or 0,
            "total_invalid_phones": row["total_invalid_phones"] or 0,
            "total_unique_phones": row["total_unique_phones"] or 0,
        }

    def get_daily_stats(self, days: int = 30) -> pd.DataFrame:
        """
        Получить статистику по дням.

        Args:
            days: количество дней назад.

        Returns:
            DataFrame с колонками: date, leads, files.
        """
        query = """
        SELECT 
            DATE(started_at) as date,
            SUM(final_rows) as leads,
            SUM(file_count) as files
        FROM processing_history
        WHERE status = 'success'
          AND started_at >= datetime('now', '-' || ? || ' days')
        GROUP BY DATE(started_at)
        ORDER BY date
        """

        rows = self.history_repo.fetch_all(query, (days,))

        if not rows:
            return pd.DataFrame(columns=["date", "leads", "files"])

        data = [{"date": row["date"], "leads": row["leads"],
                 "files": row["files"]} for row in rows]
        df = pd.DataFrame(data)
        df["date"] = pd.to_datetime(df["date"])

        return df

    def create_daily_chart(self, days: int = 30) -> Figure:
        """Создать график динамики обработки по дням (темная тема)."""
        df = self.get_daily_stats(days)

        fig, ax = plt.subplots(figsize=(14, 6))

        if df.empty:
            ax.text(
                0.5, 0.5, "Нет данных за выбранный период",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            return fig

        # Линия с градиентной заливкой
        ax.plot(
            df["date"],
            df["leads"],
            marker="o",
            linewidth=3,
            label="Лиды",
            color='#64B5F6',  # Голубой
            markerfacecolor='#42A5F5',
            markersize=8
        )
        ax.fill_between(
            df["date"],
            df["leads"],
            alpha=0.3,
            color='#64B5F6'
        )

        ax.set_xlabel("Дата", fontsize=12, color='#e0e0e0')
        ax.set_ylabel("Количество лидов", fontsize=12, color='#e0e0e0')
        ax.set_title(
            f"Динамика обработки лидов за {days} дней",
            fontsize=14,
            fontweight="bold",
            color='#ffffff',
            pad=20
        )
        ax.legend(facecolor='#2d2d2d', edgecolor='#555')
        ax.grid(True, alpha=0.2, color='#555')

        # Форматирование дат
        import matplotlib.dates as mdates
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m"))
        fig.autofmt_xdate()

        plt.tight_layout()
        return fig

    def get_manager_distribution(self, bitrix_df: pd.DataFrame) -> Dict[str, int]:
        """
        Получить распределение лидов по менеджерам.

        Args:
            bitrix_df: DataFrame с экспортированными данными.

        Returns:
            Словарь {менеджер: количество}.
        """
        if bitrix_df is None or bitrix_df.empty:
            return {}

        if "Ответственный" not in bitrix_df.columns:
            return {}

        distribution = bitrix_df["Ответственный"].value_counts().to_dict()
        return distribution

    def create_manager_pie_chart(self, bitrix_df: pd.DataFrame) -> Figure:
        """Создать круговую диаграмму распределения по менеджерам (темная тема)."""
        distribution = self.get_manager_distribution(bitrix_df)

        fig, ax = plt.subplots(figsize=(10, 10))

        if not distribution:
            ax.text(
                0.5, 0.5, "Нет данных для отображения",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            return fig

        labels = list(distribution.keys())
        sizes = list(distribution.values())

        # Темная палитра
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1',
                  '#FFA07A', '#98D8C8', '#F7DC6F']

        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            autopct="%1.1f%%",
            startangle=90,
            colors=colors,
            textprops={'color': '#e0e0e0'}
        )

        # Белый цвет для процентов
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_weight('bold')

        ax.set_title(
            "Распределение лидов по менеджерам",
            fontsize=14,
            fontweight="bold",
            color='#ffffff',
            pad=20
        )

        plt.tight_layout()
        return fig

    def create_sources_bar_chart(self, cleaned_df: pd.DataFrame) -> Figure:
        """Создать столбчатую диаграмму по источникам (темная тема)."""
        fig, ax = plt.subplots(figsize=(14, 6))

        if cleaned_df is None or cleaned_df.empty:
            ax.text(
                0.5, 0.5, "Нет данных для отображения",
                ha="center", va="center", fontsize=14, color="#999"
            )
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            return fig

        if "Источник телефона" not in cleaned_df.columns:
            ax.text(
                0.5, 0.5, "Колонка 'Источник телефона' не найдена",
                ha="center", va="center", fontsize=14, color="#999"
            )
            return fig

        sources = cleaned_df["Источник телефона"].value_counts()

        bars = ax.bar(
            range(len(sources)),
            sources.values,
            color='#9575CD',  # Фиолетовый
            edgecolor='#7E57C2',
            linewidth=1.5
        )

        # Градиент
        for i, bar in enumerate(bars):
            bar.set_color(['#9575CD', '#AB47BC', '#BA68C8',
                          '#CE93D8', '#E1BEE7'][i % 5])

        ax.set_xticks(range(len(sources)))
        ax.set_xticklabels(sources.index, rotation=45, ha="right")
        ax.set_xlabel("Источник (файл)", fontsize=12, color='#e0e0e0')
        ax.set_ylabel("Количество лидов", fontsize=12, color='#e0e0e0')
        ax.set_title(
            "Лиды по источникам",
            fontsize=14,
            fontweight="bold",
            color='#ffffff',
            pad=20
        )
        ax.grid(axis="y", alpha=0.2, color='#555')

        plt.tight_layout()
        return fig

    def export_excel_report(self, output_path: Path, stats: Dict) -> None:
        """
        Экспортировать отчёт в Excel.

        Args:
            output_path: путь для сохранения.
            stats: словарь со статистикой.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"

        # Заголовок
        ws["A1"] = "Отчёт по обработке лидов"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:B1")

        # Дата генерации
        ws["A2"] = "Дата генерации:"
        ws["B2"] = datetime.now().strftime("%d.%m.%Y %H:%M")

        # Статистика
        row = 4
        for key, value in stats.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = value
            row += 1

        # Автоширина колонок
        for col in ["A", "B"]:
            ws.column_dimensions[col].width = 30

        wb.save(output_path)
        logger.info(f"Excel-отчёт сохранён: {output_path}")
